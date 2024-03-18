import openpyxl
import ttkbootstrap as ttk
from datetime import datetime, timedelta
from ttkbootstrap.constants import *
from ttkbootstrap.scrolled import ScrolledText
from ttkbootstrap.tableview import Tableview, TableColumn
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill, Color
from ttkbootstrap.dialogs import Messagebox
from dotenv import load_dotenv
import os

load_dotenv()
path = os.getenv("CUSTOM_PATH")


class TimesheetForm(ttk.Frame):

    def __init__(self, master):
        super().__init__(master, padding=(20, 10))
        self.pack(fill=BOTH, expand=YES)
        end_date = datetime.now()
        minutes = round((end_date.minute + 7.5) / 15) * 15
        if minutes > 45:
            end_date += timedelta(hours=1)
            minutes = 0

        end_date = end_date.replace(minute=minutes)
        end_date = end_date.replace(second=0)
        start_date = end_date - timedelta(hours=1)
        fstart_date = start_date.strftime("%d/%m/%Y %H:%M:%S")
        fend_date = end_date.strftime("%d/%m/%Y %H:%M:%S")

        # form variables
        self.name = ttk.StringVar(value=os.getenv("NAME"))
        self.start = ttk.StringVar(value=fstart_date)
        self.end = ttk.StringVar(value=fend_date)
        self.detail = ttk.StringVar(value="")

        form_frame = ttk.Frame(self)
        form_frame.pack(fill=X, expand=YES)

        self.create_form_entry("NAME", self.name, form_frame)
        self.create_form_entry("START", self.start, form_frame, "date")
        self.create_form_entry("END", self.end, form_frame, "date")
        self.create_form_entry("DETAIL", self.detail, form_frame, "t_area")
        self.create_buttonbox(form_frame)

        self.create_table()

        l_frame = ttk.Frame(self)
        l_frame.pack(fill=X, pady=(15, 10))

        ref_btn = ttk.Button(
            master=l_frame,
            text="Refresh",
            command=self.refresh_table,
            bootstyle=SUCCESS,
        )
        ref_btn.pack(side=LEFT, padx=5)

        del_btn = ttk.Button(
            master=l_frame, text="Delete", command=self.on_delete, bootstyle=DANGER
        )
        del_btn.pack(side=LEFT, padx=5)

    def create_table(self):
        columns = [
            {"text": "NAME", "stretch": False},
            {"text": "START", "stretch": False},
            {"text": "END", "stretch": False},
            {"text": "DETAIL", "stretch": False},
            {"text": "MANDAYS", "stretch": False},
            {"text": "NONE", "stretch": False},
        ]

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        list_values = list(sheet.values)

        self.treeview = Tableview(
            self,
            coldata=columns,
            searchable=True,
            rowdata=list_values[1:][::-1],
            autofit=TRUE,
        )
        TableColumn(self.treeview, 0, "NAME").hide()
        TableColumn(self.treeview, 4, "MANDAYS").hide()
        TableColumn(self.treeview, 5, "NONE").hide()
        self.treeview.pack(fill=BOTH, expand=YES, padx=10, pady=(0, 7))

    def refresh_table(self, data=NONE):
        # Clear existing data in the treeview
        self.treeview.delete_rows()

        # Reload the data into the table
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        list_values = list(sheet.values)
        # last_data = list_values[-1]
        # if data is not NONE:
        #     if last_data != data:
        #         Messagebox.show_error(
        #             "Cannot submit data, please check if theres other operation opening the file"
        #         )

        self.treeview.insert_rows(0, list_values[1:][::-1])
        self.treeview.load_table_data()

    def create_form_entry(self, label, variable, container, type=None):
        container = ttk.Frame(container)
        container.pack(fill=X, expand=YES, pady=5)

        lbl = ttk.Label(master=container, text=label.title(), width=10)
        lbl.pack(side=LEFT, padx=5)

        if type == "date":
            date_object = datetime.strptime(variable.get(), "%d/%m/%Y %H:%M:%S")
            l_minutes = [0, 15, 30, 45]
            fdate = ttk.DateEntry(master=container, dateformat="%d/%m/%Y")
            fdate.pack(side=LEFT, padx=5)

            fhour = ttk.Spinbox(master=fdate, from_=9, to=19, width=3)
            fhour.pack(side=LEFT, padx=5)
            fhour.set(date_object.hour)

            fmin = ttk.Combobox(master=fdate, values=l_minutes, width=3)
            fmin.pack(side=LEFT)
            m_idx = l_minutes.index(date_object.minute)
            fmin.current(m_idx)

            fdate.bind(
                "<FocusOut>",
                lambda event, sv=variable: self.one_date_change(sv, fdate.entry.get()),
            )

            fhour.bind(
                "<FocusOut>",
                lambda event, sv=variable: self.on_hour_change(sv, fhour.get()),
            )

            fmin.bind(
                "<FocusOut>",
                lambda event, sv=variable: self.on_min_change(sv, fmin.get()),
            )

        elif type == "t_area":
            field = ScrolledText(
                master=container, height=8, width=45, wrap=WORD, autohide=TRUE
            )
            field.pack(side=LEFT, padx=3)

            field.bind(
                "<FocusOut>",
                lambda event, sv=variable: self.on_change(
                    sv, field.get("1.0", "end-1c")
                ),
            )

        else:
            field = ttk.Entry(master=container, textvariable=variable)
            field.pack(side=LEFT, padx=5, fill=X)

    def on_change(self, variable, new_value):
        if variable.get() != new_value:
            variable.set(new_value)

    def one_date_change(self, variable, new_value):
        new_date = datetime.strptime(new_value, "%d/%m/%Y")
        date_obj = datetime.strptime(variable.get(), "%d/%m/%Y %H:%M:%S")
        updated_date_obj = date_obj.replace(
            day=new_date.day, month=new_date.month, year=new_date.year
        )
        variable.set(updated_date_obj.strftime("%d/%m/%Y %H:%M:%S"))

    def on_hour_change(self, variable, new_value):
        date_obj = datetime.strptime(variable.get(), "%d/%m/%Y %H:%M:%S")
        updated_date_obj = date_obj.replace(hour=int(new_value))
        variable.set(updated_date_obj.strftime("%d/%m/%Y %H:%M:%S"))

    def on_min_change(self, variable, new_value):
        date_obj = datetime.strptime(variable.get(), "%d/%m/%Y %H:%M:%S")
        updated_date_obj = date_obj.replace(minute=int(new_value))
        variable.set(updated_date_obj.strftime("%d/%m/%Y %H:%M:%S"))

    def create_buttonbox(self, container):
        container = ttk.Frame(container)
        container.pack(fill=X, expand=YES, pady=(15, 10))

        sub_btn = ttk.Button(
            master=container,
            text="Submit",
            command=self.on_submit,
            bootstyle=SUCCESS,
            width=6,
        )
        sub_btn.pack(side=LEFT, padx=5)
        sub_btn.focus_set()

        cnl_btn = ttk.Button(
            master=container,
            text="Cancel",
            command=self.on_cancel,
            bootstyle=DANGER,
            width=6,
        )
        cnl_btn.pack(side=LEFT, padx=5)

    def on_submit(self):

        name = self.name.get()
        start = self.start.get()
        end = self.end.get()
        detail = self.detail.get()

        start = datetime.strptime(start, "%d/%m/%Y %H:%M:%S")
        start_formatted = start.strftime("%d-%b-%y %H.%M")

        end = datetime.strptime(end, "%d/%m/%Y %H:%M:%S")
        end_formatted = end.strftime("%d-%b-%y %H.%M")

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        alignment_style = Alignment(wrapText=TRUE, horizontal=CENTER, vertical=CENTER)

        font_style = Font(size="10")

        new_row_index = sheet.max_row + 1
        name_cell = sheet.cell(row=new_row_index, column=1)
        name_cell.value = name
        name_cell.border = border_style
        name_cell.alignment = alignment_style
        name_cell.font = font_style

        start_cell = sheet.cell(row=new_row_index, column=2)
        start_cell.value = start_formatted
        start_cell.border = border_style
        start_cell.alignment = alignment_style
        start_cell.font = font_style
        start_cell.fill = PatternFill(
            patternType="solid", fill_type="solid", fgColor=Color("f8e5d8")
        )

        end_cell = sheet.cell(row=new_row_index, column=3)
        end_cell.value = end_formatted
        end_cell.border = border_style
        end_cell.alignment = alignment_style
        end_cell.font = font_style
        end_cell.fill = PatternFill(
            patternType="solid", fill_type="solid", fgColor=Color("f8e5d8")
        )

        detail_cell = sheet.cell(row=new_row_index, column=4)
        detail_cell.value = detail
        detail_cell.border = border_style
        detail_cell.alignment = alignment_style
        detail_cell.font = font_style

        mandays_cell = sheet.cell(row=new_row_index, column=5)
        mandays_cell.value = (
            f"=ROUND((HOUR(${end_cell.coordinate}-${start_cell.coordinate})/8), 2)"
        )
        mandays_cell.border = border_style
        mandays_cell.alignment = alignment_style
        mandays_cell.font = font_style
        mandays_cell.fill = PatternFill(
            patternType="solid", fill_type="solid", fgColor=Color("e4efdc")
        )

        workbook.save(path)

        t_data = (
            name,
            start_formatted,
            end_formatted,
            detail,
            mandays_cell.value,
        )

        self.refresh_table(t_data)

    def on_cancel(self):
        self.quit()

    def on_delete(self):
        d_values = []
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        s_rows = self.treeview.view.selection()
        for row_id in s_rows:
            value = self.treeview.view.item(row_id, "values")
            for row in sheet.iter_rows(values_only=True):
                t_row = tuple("None" if value is None else value for value in row)
                if t_row == value:
                    d_values.append(t_row)
                    index = list(sheet.iter_rows(values_only=True)).index(row)
                    sheet.delete_rows(index + 1, 1)
            self.treeview.delete_row(iid=row_id)
        workbook.save(path)
        list_values = list(sheet.values)
        for value in d_values:
            if value in list_values:
                Messagebox.show_error(
                    "Cannot delete data, please check if theres other operation opening the file"
                )

        self.treeview.load_table_data()


if __name__ == "__main__":
    app = ttk.Window("Timesheet", "solar", resizable=(True, False))
    timesheet = TimesheetForm(app)
    timesheet.pack(padx=10, pady=10)
    app.mainloop()
